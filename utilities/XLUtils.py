import openpyxl

def getRowCount(file,sheetname):
    excelfile=openpyxl.load_workbook(file)
    sheet=excelfile[sheetname]
    return (sheet.max_row)

def getColCount(file,sheetname):
    excelfile=openpyxl.load_workbook(file)
    sheet=excelfile[sheetname]
    return (sheet.max_column)

def ReadData(file, sheetname, rowno, colno):
    excelfile=openpyxl.load_workbook(file)
    sheet=excelfile[sheetname]
    return (sheet.cell(row=rowno,column=colno)).value

def WriteData(file, sheetname, rowno, colno, data):
    excelfile=openpyxl.load_workbook(file)
    sheet=excelfile[sheetname]
    sheet.cell(row=rowno,column=colno).value=data